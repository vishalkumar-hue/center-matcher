import pandas as pd
import re
from difflib import SequenceMatcher
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

MATCH_THRESHOLD = 65
COL_NAME = "TEST CENTER NAME"

DISTRICT_ALIASES = {
    "KANPUR NAGAR": "KANPUR",
    "KANPUR DEHAT": "KANPUR DEHAT",
}
DISTRICT_FUZZY_THRESHOLD = 0.82

# ---- FIX 2: gic/ggic ab decompose hote hain "govt ic" / "govt girls ic" mein,
# taaki "G.I.C." abbreviation aur "Government Inter College" full-form
# hamesha SAME normalized tokens produce karein, chahe original text mein
# jo bhi form likha ho.
CANONICAL = [
    (r'\bgovernment\s+inter\s+col(?:lege)?\b',       'govt ic'),
    (r'\bgovt\.?\s+inter\s+col(?:lege)?\b',          'govt ic'),
    (r'\braj(?:kiya|keeya|akiya|akeeya|keya|kiyaa|kiye|kiyay)\s+inter\s+col(?:lege)?\b', 'govt ic'),
    (r'\bpm\s+shri\s+(?:govt|government|rajkiya)\b', 'govt'),
    (r'\bg\s*g\s*i\s*c\b',                           'govt girls ic'),   # GGIC (merged acronym)
    (r'\bg\s*i\s*c\b',                                'govt ic'),        # GIC (merged acronym)
    (r'\bpost\s*gradu(?:ate|ation)\s*col(?:lege)?\b', 'pgcollege'),
    (r'\bdegree\s*col(?:lege)?\b',                    'pgcollege'),
    (r'\bmahavidyalaya\b',                            'pgcollege'),
    (r'\bsnatakottar\s*(?:col(?:lege)?)?\b',          'pgcollege'),
    (r'\bsnatakor\s*(?:col(?:lege)?)?\b',             'pgcollege'),
    (r'\bpg\s*col(?:lege)?\b',                        'pgcollege'),
    (r'\binter(?:mediate)?\s*col(?:lege)?\b',         'ic'),
    (r'\bgovernment\b',  'govt'),
    (r'\bgovt\.?\b',     'govt'),
    (r'\braj(?:kiya|keeya|akiya|akeeya|keya|kiyaa|kiye|kiyay)\b', 'govt'),
    (r'\bbalika\b',  'girls'),
    (r'\bkanya\b',   'girls'),
    (r'\bkumari\b',  'girls'),
    (r'\bmahila\b',  'girls'),
    (r'\bwomens?\b', 'girls'),
    (r'\bbalak\b',   'boys'),
    (r'\bvidyalayam?\b', 'school'),
    (r'\bvidyalay\b',    'school'),
    (r'\bd\s*a\s*v\b',  'dav'),
    (r'\bpost\s*gradu(?:ate|ation)\b',  'pg'),
    (r'\bp\s*g\b',               'pg'),
    (r'\b(?:senior|higher)\s*secondary\b', 'srsc'),
    (r'\bsr\.?\s*sec\.?\b',               'srsc'),
    (r'\bss\s*school\b',                  'srsc'),
    (r'\bf\s*e\s*t\b',  'fet'),
    (r'\badarsh\b',  'ideal'),
    (r'\bsmarak\b',  'memorial'),
    (r'\bsmriti\b',  'memorial'),
    (r'\bpublic\s*school\b',  'ps'),
    (r'\bvidyapeeth\b',   'vidyapith'),
    (r'\bvidyapeetha\b',  'vidyapith'),
    (r'\bshree\b',  'shri'),
    (r'\bsri\b',    'shri'),
    (r'\bkendriya\s*school\b',   'kv'),
    (r'\bkendriya\s*vidyalya\b', 'kv'),
    (r'\bkv\b',                  'kv'),
    (r'\bhindi\s*medium\b',   ''),
    (r'\benglish\s*medium\b', ''),
    (r'\bco\s*ed\b',          ''),
    (r'\bpm\s+shri\b',        ''),
    (r'\bautonomous\b',       ''),
    # FIX 4: "+2" / "plus two" -> ek hi token
    (r'\bplus\s*two\b', 'p2'),
    (r'\+\s*2\b',       'p2'),
]

NOISE = {
    'the', 'and', 'of', 'for', 'in', 'at', 'by', 'near', 'new', 'old',
    'no', 'number', 'sh', 'smt', 'dr', 'km', 'late', 'pt', 'prof',
    'agra', 'lucknow', 'kanpur', 'varanasi', 'prayagraj', 'aligarh',
    'ayodhya', 'azamgarh', 'mathura', 'bareilly', 'moradabad', 'meerut',
    'ghaziabad', 'noida', 'gorakhpur', 'allahabad', 'jhansi', 'firozabad',
    'basti', 'saharanpur', 'mirzapur', 'banda',
}

STRUCTURAL_TOKENS = {'ic', 'govt', 'pgcollege', 'pg', 'srsc', 'ps',
                     'kv', 'dav', 'school', 'college', 'girls', 'boys', 'shri', 'ideal',
                     'memorial', 'vidyapith', 'fet', 'women', 'p2',
                     'blka', 'blkb', 'blkc', 'blkd', 'blke', 'blkf',
                     'blk1', 'blk2', 'blk3', 'blk4', 'blk5'}

# FIX 6: A/B/C style block labels aur 1/2/3 style dono ek hi cheez hote
# hain (List A "BLOCK-A" likhta hai, List B "BLOCK-1" likhta hai). Isse
# normalize karke same token pe le aate hain taaki block-match bonus
# dono naming scheme mein kaam kare.
BLOCK_ALIAS = {'blka': 'blk1', 'blkb': 'blk2', 'blkc': 'blk3',
               'blkd': 'blk4', 'blke': 'blk5', 'blkf': 'blk6'}


def _normalize_block_tokens(words):
    return [BLOCK_ALIAS.get(w, w) for w in words]

# FIX 3: Fuzzy vocabulary snap -- common data-entry typos jaise
# "mahavidyala", "colege", "uccha" ko unke sahi canonical spelling
# se map kar deta hai, taaki CANONICAL phrase regex unhe pakad sake.
# Ye endless "har typo ke liye naya regex" likhne se behtar hai --
# generalize karta hai kisi bhi chhoti spelling mistake ke liye.
VOCAB_WORDS = [
    'mahavidyalaya', 'vidyalaya', 'uchcha', 'college', 'inter',
    'intermediate', 'secondary', 'vidyapeeth', 'snatakottar', 'vidyapith',
    'government', 'rajkiya', 'balika', 'kanya', 'mahila', 'smarak',
    'adarsh', 'public', 'school', 'vidyapeetha', 'mahavidyala',
]
VOCAB_FUZZY_THRESHOLD = 0.82


def _fuzzy_snap_word(w):
    if len(w) < 5:
        return w
    best, best_score = None, 0.0
    for v in VOCAB_WORDS:
        if abs(len(v) - len(w)) > 3:
            continue
        score = SequenceMatcher(None, w, v).ratio()
        if score > best_score:
            best_score, best = score, v
    if best_score >= VOCAB_FUZZY_THRESHOLD and best != w:
        # 'mahavidyala' khud VOCAB mein hai as an alias -> asli target 'mahavidyalaya'
        return 'mahavidyalaya' if best == 'mahavidyala' else best
    return w


# ---- FIX 1: acronym merge ab dot AND space dono se separated single-letter
# tokens ko merge karta hai (e.g. "N A S" ya "N.A.S." dono "nas" ban jaate
# hain). Pehle sirf dotted version handle hoti thi, isliye "N A S COLLEGE"
# (bina dot ke) kabhi merge nahi hota tha aur match fail ho jaata tha.
def _merge_acronyms(x):
    x = re.sub(r'\.', '. ', x)
    x = re.sub(r'\s+', ' ', x).strip()
    tokens = x.split(' ') if x else []
    out = []
    buf = []

    def flush():
        if len(buf) >= 2:
            out.append(''.join(buf))
        elif len(buf) == 1:
            out.append(buf[0])
        buf.clear()

    for tok in tokens:
        core = tok.rstrip('.')
        if len(core) == 1 and core.isalpha():
            buf.append(core)
        else:
            flush()
            out.append(core if tok.endswith('.') else tok)
    flush()
    return ' '.join(out)


def clean_text(x):
    x = str(x).lower()
    x = re.sub(r'\(?\s*block[-\s]*([a-z0-9]+)\s*\)?', r' blk\1 ', x)
    x = _merge_acronyms(x)
    x = re.sub(r'[^a-z0-9 ]', ' ', x)
    x = re.sub(r'\s+', ' ', x).strip()
    # fuzzy-snap typo'd words BEFORE phrase-level canonical regex runs
    x = ' '.join(_fuzzy_snap_word(w) for w in x.split())
    for pattern, replacement in CANONICAL:
        x = re.sub(pattern, replacement, x)
    x = re.sub(r'\s+', ' ', x).strip()
    words = [w for w in x.split() if w not in NOISE and len(w) > 0]
    words = _normalize_block_tokens(words)
    return " ".join(words).strip()


WORD_FUZZY_THRESHOLD = 0.82


def _fuzzy_content_common_count(wa, wb):
    content_a = [w for w in wa if len(w) > 2 and w not in STRUCTURAL_TOKENS]
    content_b = [w for w in wb if len(w) > 2 and w not in STRUCTURAL_TOKENS]
    used_b = set()
    count = 0
    for w in content_a:
        if w in content_b and w not in used_b:
            used_b.add(w)
            count += 1
            continue
        best, best_score = None, 0.0
        for wb_word in content_b:
            if wb_word in used_b:
                continue
            score = SequenceMatcher(None, w, wb_word).ratio()
            if score > best_score:
                best_score, best = score, wb_word
        if best_score >= WORD_FUZZY_THRESHOLD:
            used_b.add(best)
            count += 1
    return count


def _containment_score(a_words, b_words):
    """
    FIX 5: Jab ek naam doosre ke andar word-sequence ke roop mein
    poora contained hai (e.g. "gic" vs "gic bahraich"), purana formula
    (character-level SequenceMatcher ratio) length difference ki wajah se
    bahut kam score deta tha -- chahe match 100% sahi ho. Ye naya word-level
    containment check us case ko fairly zyada score deta hai.
    """
    if not a_words or not b_words:
        return 0
    short, long_ = (a_words, b_words) if len(a_words) <= len(b_words) else (b_words, a_words)
    # check short is a subsequence of long_ (order preserved)
    it = iter(long_)
    if all(w in it for w in short):
        ratio = len(short) / len(long_)
        return int(50 + 50 * ratio)
    return 0


# FIX 8: Bahut saare college naam ek taraf poora likhe hote hain
# ("SHRI SANATAN DHARAM", "RAJA BALWANT SINGH") aur doosri taraf sirf
# initials ke roop mein ("SSD", "RBS"). Ye koi typo nahi hai -- ye ek
# legitimate short-form hai jo bahut common hai UP govt college naamo
# mein. Ye function check karta hai ki koi chhota acronym-jaisa word
# (2-6 letters) doosre naam ke kisi lagatar (contiguous) words ke
# pehle-akshar se match karta hai ya nahi.
def _initials_match(a_words, b_words):
    def check(acronym_side, full_side):
        for w in acronym_side:
            if not (2 <= len(w) <= 6) or not w.isalpha() or w in STRUCTURAL_TOKENS:
                continue
            n = len(w)
            for i in range(len(full_side) - n + 1):
                window = full_side[i:i + n]
                if any(len(t) == 0 for t in window):
                    continue
                initials = ''.join(t[0] for t in window)
                if initials == w:
                    return True
        return False
    return check(a_words, b_words) or check(b_words, a_words)


def similarity(a, b):
    if not a or not b:
        return 0
    s1 = int(SequenceMatcher(None, a, b).ratio() * 100)
    s2 = int(SequenceMatcher(None, " ".join(sorted(a.split())), " ".join(sorted(b.split()))).ratio() * 100)
    a_words = a.split()
    b_words = b.split()
    content_a_n = len([w for w in a_words if len(w) > 2 and w not in STRUCTURAL_TOKENS])
    content_b_n = len([w for w in b_words if len(w) > 2 and w not in STRUCTURAL_TOKENS])
    content_common_count = _fuzzy_content_common_count(a_words, b_words)
    s3 = 0
    if content_common_count:
        # FIX 7: denominator ab sirf CONTENT words ka count hai (structural
        # tokens jaise 'ic'/'govt'/'college' exclude), na ki poore words ka.
        # Isse abbreviation-vs-full-name pairs jaise "R B S College" vs
        # "Raja Balwant Singh (RBS) Inter College" sahi se match hote hain --
        # pehle poore word count denominator mein hone se score bahut dilute
        # ho jaata tha.
        shorter = min(content_a_n, content_b_n)
        if shorter > 0:
            s3 = int(content_common_count / shorter * 100)
    s4 = 0
    if len(a.split()) <= 2 or len(b.split()) <= 2:
        short = a if len(a) <= len(b) else b
        long_ = b if len(a) <= len(b) else a
        if short and short in long_:
            s4 = int(SequenceMatcher(None, short, long_).ratio() * 100)
    s5 = _containment_score(a.split(), b.split())
    raw = max(s1, s2, s3, s4, s5)
    if raw < 80 and _initials_match(a_words, b_words):
        raw = max(raw, 80)
    blk_a = {w for w in a.split() if w.startswith('blk')}
    blk_b = {w for w in b.split() if w.startswith('blk')}
    if blk_a and blk_b:
        if blk_a == blk_b:
            raw = min(100, raw + 5)
        else:
            raw = min(raw, 55)
    return raw


def normalize_district(d):
    d = str(d).upper().strip()
    d = re.sub(r'[.\-_/,()]', ' ', d)
    d = re.sub(r'\bDISTT\.?\b', '', d)
    d = re.sub(r'\bDISTRICT\b', '', d)
    d = re.sub(r'\s+', ' ', d).strip()
    return DISTRICT_ALIASES.get(d, d)


class DistrictMatcher:
    """
    Exact district match na milne par closest existing district
    (B side) dhoondh leta hai, taaki chhoti si spelling/spacing
    difference se pura match zero na ho jaaye. Result cache hota
    hai taaki baar baar recompute na ho.
    """
    def __init__(self, grouped, threshold=DISTRICT_FUZZY_THRESHOLD):
        self.grouped = grouped
        self.threshold = threshold
        self.cache = {}
        self.unmatched_districts = set()

    def get_group(self, dist):
        if dist in self.grouped.groups:
            return self.grouped.get_group(dist)
        if dist in self.cache:
            cand = self.cache[dist]
            return self.grouped.get_group(cand) if cand else None
        best, best_score = None, 0.0
        for cand in self.grouped.groups.keys():
            score = SequenceMatcher(None, dist, cand).ratio()
            if score > best_score:
                best_score, best = score, cand
        if best_score >= self.threshold:
            self.cache[dist] = best
            return self.grouped.get_group(best)
        self.cache[dist] = None
        self.unmatched_districts.add(dist)
        return None


# ═══════════════════════════════════════════════════════════════
#  SINGLE B FILE MATCHING (purana, single tab output)
# ═══════════════════════════════════════════════════════════════
def run_matching(file_a_path, file_b_path, output_path):
    df_a = pd.read_excel(file_a_path)
    df_b = pd.read_excel(file_b_path)
    df_a.columns = df_a.columns.str.strip()
    df_b.columns = df_b.columns.str.strip()

    COL_A_DIST = [c for c in df_a.columns if 'district' in c.lower()][0]
    COL_B_DIST = [c for c in df_b.columns if 'district' in c.lower()][0]

    b_state_col = [c for c in df_b.columns if 'state' in c.lower()]
    if b_state_col:
        df_b_filtered = df_b[df_b[b_state_col[0]].str.upper().str.contains('UTTAR PRADESH', na=False)].copy()
    else:
        df_b_filtered = df_b.copy()

    df_a["_name"] = df_a[COL_NAME].apply(clean_text)
    df_b_filtered["_name"] = df_b_filtered[COL_NAME].apply(clean_text)
    df_a["_dist"] = df_a[COL_A_DIST].apply(normalize_district)
    df_b_filtered["_dist"] = df_b_filtered[COL_B_DIST].apply(normalize_district)

    b_grouped = df_b_filtered.groupby("_dist")
    dmatch = DistrictMatcher(b_grouped)

    candidates_all = []
    for idx_a, row_a in df_a.iterrows():
        group = dmatch.get_group(row_a["_dist"])
        if group is not None:
            for idx_b, row_b in group.iterrows():
                score = similarity(row_a["_name"], row_b["_name"])
                if score >= MATCH_THRESHOLD:
                    candidates_all.append((score, idx_a, idx_b))

    candidates_all.sort(key=lambda x: -x[0])
    used_a, used_b, matches = set(), set(), {}
    for score, idx_a, idx_b in candidates_all:
        if idx_a not in used_a and idx_b not in used_b:
            matches[idx_a] = (idx_b, score)
            used_a.add(idx_a)
            used_b.add(idx_b)

    matched_rows = []
    for idx_a, row_a in df_a.iterrows():
        a_cols = list(row_a.drop(["_name", "_dist"]))
        if idx_a in matches:
            idx_b, score = matches[idx_a]
            b_cols = list(df_b_filtered.loc[idx_b].drop(["_name", "_dist"]))
            matched_rows.append(a_cols + b_cols + [score, "MATCHED"])
        else:
            best_score = 0
            group = dmatch.get_group(row_a["_dist"])
            if group is not None:
                for idx_b, row_b in group.iterrows():
                    s = similarity(row_a["_name"], row_b["_name"])
                    if s > best_score:
                        best_score = s
            b_cols = [None] * (len(df_b_filtered.columns) - 2)
            matched_rows.append(a_cols + b_cols + [best_score, "UNMATCHED_A"])

    for idx_b, row_b in df_b_filtered.iterrows():
        if idx_b not in used_b:
            a_cols = [None] * (len(df_a.columns) - 2)
            b_cols = list(row_b.drop(["_name", "_dist"]))
            matched_rows.append(a_cols + b_cols + [0, "UNMATCHED_B"])

    a_final_cols = list(df_a.columns.drop(["_name", "_dist"]))
    b_final_cols = list(df_b_filtered.columns.drop(["_name", "_dist"]))
    b_renamed = ["B_" + c if c in a_final_cols else c for c in b_final_cols]
    columns = a_final_cols + b_renamed + ["Match %", "Status"]
    final_df = pd.DataFrame(matched_rows, columns=columns)

    matched_count   = len(final_df[final_df["Status"] == "MATCHED"])
    unmatched_a_cnt = len(final_df[final_df["Status"] == "UNMATCHED_A"])
    unmatched_b_cnt = len(final_df[final_df["Status"] == "UNMATCHED_B"])
    match_rate      = round(matched_count / max(len(df_a), 1) * 100, 1)

    final_df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active
    GREEN  = PatternFill("solid", fgColor="C6EFCE")
    RED    = PatternFill("solid", fgColor="FFC7CE")
    YELLOW = PatternFill("solid", fgColor="FFEB9C")
    HEADER = PatternFill("solid", fgColor="4472C4")
    WHITE_FONT = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = HEADER
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    status_col = columns.index("Status") + 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        status = row[status_col - 1].value
        fill = GREEN if status == "MATCHED" else (RED if status == "UNMATCHED_A" else YELLOW)
        for cell in row:
            cell.fill = fill
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)
    ws.freeze_panes = "A2"
    wb.save(output_path)

    return {
        "total": len(final_df),
        "matched": matched_count,
        "unmatched_a": unmatched_a_cnt,
        "unmatched_b": unmatched_b_cnt,
        "match_rate": match_rate,
        "unmatched_districts": sorted(dmatch.unmatched_districts),
    }


# ═══════════════════════════════════════════════════════════════
#  MULTI B FILE MATCHING (naya, multi-tab output)
# ═══════════════════════════════════════════════════════════════
def run_matching_multi(file_a_path, file_b_paths_with_names, output_path):
    """
    file_a_path: List A ka path (string)
    file_b_paths_with_names: list of tuples [(tab_name, file_path), ...]
    output_path: output excel ka path
    """
    df_a = pd.read_excel(file_a_path)
    df_a.columns = df_a.columns.str.strip()

    COL_A_DIST = [c for c in df_a.columns if 'district' in c.lower()][0]
    COL_A_STATE = [c for c in df_a.columns if 'state' in c.lower()]

    df_a["_name"] = df_a[COL_NAME].apply(clean_text)
    df_a["_dist"] = df_a[COL_A_DIST].apply(normalize_district)
    if COL_A_STATE:
        df_a["_state"] = df_a[COL_A_STATE[0]].str.upper().str.strip()
    else:
        df_a["_state"] = ""

    sheet_results = {}
    per_file_stats = []
    all_unmatched_districts = set()

    for tab_name, file_b_path in file_b_paths_with_names:
        df_b = pd.read_excel(file_b_path)
        df_b.columns = df_b.columns.str.strip()

        COL_B_DIST = [c for c in df_b.columns if 'district' in c.lower()][0]
        b_state_col = [c for c in df_b.columns if 'state' in c.lower()]

        df_b["_name"] = df_b[COL_NAME].apply(clean_text)
        df_b["_dist"] = df_b[COL_B_DIST].apply(normalize_district)

        if b_state_col:
            b_state = str(df_b[b_state_col[0]].iloc[0]).upper().strip()
            df_a_f = df_a[df_a["_state"] == b_state].copy() if COL_A_STATE else df_a.copy()
        else:
            df_a_f = df_a.copy()

        if df_a_f.empty:
            a_raw_cols = [c for c in df_a.columns if not c.startswith('_')]
            b_raw_cols = [c for c in df_b.columns if not c.startswith('_')]
            b_renamed = ["B_" + c if c in a_raw_cols else c for c in b_raw_cols]
            rows = []
            for _, row_b in df_b.iterrows():
                b_cols = list(row_b[b_raw_cols])
                rows.append([None] * len(a_raw_cols) + b_cols + [0, "UNMATCHED_B"])
            cols = a_raw_cols + b_renamed + ["Match %", "Status"]
            result_df = pd.DataFrame(rows, columns=cols)
            sheet_results[tab_name] = result_df
            per_file_stats.append({
                "tab": tab_name, "matched": 0, "unmatched_a": 0,
                "unmatched_b": len(df_b), "match_rate": 0.0
            })
            continue

        b_grouped = df_b.groupby("_dist")
        dmatch = DistrictMatcher(b_grouped)

        candidates_all = []
        for idx_a, row_a in df_a_f.iterrows():
            group = dmatch.get_group(row_a["_dist"])
            if group is not None:
                for idx_b, row_b in group.iterrows():
                    score = similarity(row_a["_name"], row_b["_name"])
                    if score >= MATCH_THRESHOLD:
                        candidates_all.append((score, idx_a, idx_b))

        candidates_all.sort(key=lambda x: -x[0])
        used_a, used_b, matches = set(), set(), {}
        for score, idx_a, idx_b in candidates_all:
            if idx_a not in used_a and idx_b not in used_b:
                matches[idx_a] = (idx_b, score)
                used_a.add(idx_a)
                used_b.add(idx_b)

        a_raw_cols = [c for c in df_a_f.columns if not c.startswith('_')]
        b_raw_cols = [c for c in df_b.columns if not c.startswith('_')]
        b_renamed = ["B_" + c if c in a_raw_cols else c for c in b_raw_cols]

        rows = []
        for idx_a, row_a in df_a_f.iterrows():
            a_cols = list(row_a[a_raw_cols])
            if idx_a in matches:
                idx_b, score = matches[idx_a]
                b_cols = list(df_b.loc[idx_b, b_raw_cols])
                rows.append(a_cols + b_cols + [score, "MATCHED"])
            else:
                best_score = 0
                group = dmatch.get_group(row_a["_dist"])
                if group is not None:
                    for idx_b, row_b in group.iterrows():
                        s = similarity(row_a["_name"], row_b["_name"])
                        if s > best_score:
                            best_score = s
                rows.append(a_cols + [None] * len(b_raw_cols) + [best_score, "UNMATCHED_A"])

        for idx_b, row_b in df_b.iterrows():
            if idx_b not in used_b:
                rows.append([None] * len(a_raw_cols) + list(row_b[b_raw_cols]) + [0, "UNMATCHED_B"])

        cols = a_raw_cols + b_renamed + ["Match %", "Status"]
        result_df = pd.DataFrame(rows, columns=cols)
        sheet_results[tab_name] = result_df

        m  = (result_df["Status"] == "MATCHED").sum()
        ua = (result_df["Status"] == "UNMATCHED_A").sum()
        ub = (result_df["Status"] == "UNMATCHED_B").sum()
        per_file_stats.append({
            "tab": tab_name, "matched": int(m), "unmatched_a": int(ua),
            "unmatched_b": int(ub),
            "match_rate": round(m / max(m + ua, 1) * 100, 1)
        })
        all_unmatched_districts |= dmatch.unmatched_districts

    GREEN  = PatternFill("solid", fgColor="C6EFCE")
    RED    = PatternFill("solid", fgColor="FFC7CE")
    YELLOW = PatternFill("solid", fgColor="FFEB9C")
    HEADER = PatternFill("solid", fgColor="4472C4")
    WHITE_FONT = Font(bold=True, color="FFFFFF")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df = pd.DataFrame(per_file_stats)
        summary_df = summary_df[["tab", "matched", "unmatched_a", "unmatched_b", "match_rate"]]
        summary_df.columns = ["Tab", "Matched", "Unmatched A", "Unmatched B", "Match Rate %"]
        summary_df.to_excel(writer, sheet_name="SUMMARY", index=False)

        for tab_name, df in sheet_results.items():
            df.to_excel(writer, sheet_name=tab_name[:31], index=False)

    wb = load_workbook(output_path)

    def format_ws(ws, has_status=True):
        for cell in ws[1]:
            cell.fill = HEADER
            cell.font = WHITE_FONT
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.freeze_panes = "A2"
        if has_status:
            status_col = next((i for i, c in enumerate(ws[1], 1) if c.value == "Status"), None)
            if status_col:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    s = row[status_col - 1].value
                    fill = GREEN if s == "MATCHED" else (RED if s == "UNMATCHED_A" else YELLOW)
                    for cell in row:
                        cell.fill = fill
        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = max((len(str(c.value)) if c.value else 0) for c in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    format_ws(wb["SUMMARY"], has_status=False)
    for tab_name in sheet_results:
        sheet_name = tab_name[:31]
        if sheet_name in wb.sheetnames:
            format_ws(wb[sheet_name])
    wb.save(output_path)

    total_matched = sum(s["matched"] for s in per_file_stats)
    total_unmatched_a = sum(s["unmatched_a"] for s in per_file_stats)
    total_unmatched_b = sum(s["unmatched_b"] for s in per_file_stats)

    return {
        "per_file": per_file_stats,
        "total_matched": total_matched,
        "total_unmatched_a": total_unmatched_a,
        "total_unmatched_b": total_unmatched_b,
        "unmatched_districts": sorted(all_unmatched_districts),
    }
